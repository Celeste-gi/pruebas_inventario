# -*- coding: utf-8 -*-
"""
Created on Thu Nov  7 01:56:26 2024

@author: celes
"""


import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import pandas as pd
from PIL import Image, ImageTk  # Librerías para cargar la imagen del logo

# Crear la ventana principal de Tkinter
ventana = tk.Tk()
ventana.title("Inventario de Refrigerador -20°")
ventana.geometry("800x700")
ventana.config(bg="#F0E6F7")  # Fondo de la ventana en tono lavanda suave

# Imagen del logo del laboratorio
try:
    logo_path = r"C:\Users\celes\pruebas_inventario\logo.jpg"  # Ruta de la imagen del logo
    logo_image = Image.open(logo_path)
    logo_image = logo_image.resize((800, 100), Image.Resampling.LANCZOS)  # Ajustar el tamaño
    logo_photo = ImageTk.PhotoImage(logo_image)
    logo_label = tk.Label(ventana, image=logo_photo, bg="#F0E6F7")
    logo_label.grid(row=0, column=0, columnspan=2, pady=(20, 30))
except FileNotFoundError:
    messagebox.showwarning("Advertencia", "No se encontró el logo en la ruta especificada.")

# Variables globales
ruta_excel = None
wb = None
hojas = []

# Función para seleccionar archivo Excel
def seleccionar_archivo():
    global ruta_excel, wb, hojas
    ruta_excel = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
    if ruta_excel:
        try:
            wb = load_workbook(ruta_excel)
            hojas = wb.sheetnames
            hoja_combobox["values"] = hojas
            hoja_combobox.current(0)
            messagebox.showinfo("Archivo seleccionado", f"Archivo seleccionado: {ruta_excel}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar el archivo: {e}")

# Crear botón para seleccionar archivo
boton_seleccionar_archivo = tk.Button(ventana, text="Seleccionar Archivo Excel", command=seleccionar_archivo, bg="#F3A5F3", fg="white", font=("Arial", 10, "bold"))
boton_seleccionar_archivo.grid(row=1, column=0, columnspan=2, pady=10)

# Campos de entrada para cada columna
campos = {
    "Contenedor": tk.StringVar(),
    "Etiqueta": tk.StringVar(),
    "Nombre": tk.StringVar(),
    "Subtitulo": tk.StringVar(),
    "Marca": tk.StringVar(),
    "F. vencimiento": tk.StringVar(),  # Formato: YYYY-MM-DD
    "Detalle": tk.StringVar(),
    "Unidades": tk.IntVar()
}

# Crear un combobox para seleccionar la hoja de trabajo
tk.Label(ventana, text="Seleccionar Bandeja", bg="#F0E6F7", fg="#333", font=("Arial", 10, "bold")).grid(row=2, column=0, padx=10, pady=5, sticky="w")
bandeja_seleccionada = tk.StringVar()
hoja_combobox = ttk.Combobox(ventana, textvariable=bandeja_seleccionada, state="readonly")
hoja_combobox.grid(row=2, column=1, padx=10, pady=5)

# Función para actualizar el stock (añadir o restar unidades)
def actualizar_stock(modo):
    if not wb or not bandeja_seleccionada.get():
        messagebox.showwarning("Advertencia", "Primero selecciona un archivo y una hoja.")
        return
    
    hoja = wb[bandeja_seleccionada.get()]
    datos = {columna: variable.get() for columna, variable in campos.items()}
    
    if datos["F. vencimiento"]:
        try:
            datetime.strptime(datos["F. vencimiento"], "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error", "Fecha de vencimiento no válida. Formato correcto: YYYY-MM-DD.")
            return

    nombre_columna, unidades_columna = None, None
    columnas_hoja = [cell.value for cell in hoja[1]]
    
    for cell in hoja[1]:
        if cell.value == "Nombre":
            nombre_columna = cell.column
        elif cell.value == "Unidades":
            unidades_columna = cell.column

    if not nombre_columna or not unidades_columna:
        messagebox.showerror("Error", "No se encontraron las columnas 'Nombre' o 'Unidades' en la hoja seleccionada.")
        return

    fila_producto = None
    for fila in range(2, hoja.max_row + 1):
        if hoja.cell(row=fila, column=nombre_columna).value == datos["Nombre"]:
            fila_producto = fila
            break

    if fila_producto:
        unidades_actuales = hoja.cell(row=fila_producto, column=unidades_columna).value or 0
        if modo == "añadir":
            hoja.cell(row=fila_producto, column=unidades_columna, value=unidades_actuales + datos["Unidades"])
        elif modo == "restar":
            nueva_unidades = max(0, unidades_actuales - datos["Unidades"])
            hoja.cell(row=fila_producto, column=unidades_columna, value=nueva_unidades)
        
        for idx, (columna, valor) in enumerate(datos.items(), start=1):
            if columna != "Unidades" and valor and valor != hoja.cell(row=fila_producto, column=idx).value:
                messagebox.showwarning("Advertencia", f"Contradicción en la columna '{columna}' para el producto '{datos['Nombre']}'.")
                return
    else:
        ultima_fila = hoja.max_row + 1
        for idx, (columna, valor) in enumerate(datos.items(), start=1):
            hoja.cell(row=ultima_fila, column=idx, value=valor)

    wb.save(ruta_excel)
    messagebox.showinfo("Éxito", f"Producto {'añadido' if modo == 'añadir' else 'restado'} correctamente.")
    for variable in campos.values():
        if isinstance(variable, tk.StringVar):
            variable.set("")
        elif isinstance(variable, tk.IntVar):
            variable.set(0)

# Función para mostrar los productos en una tabla
def ver_articulos():
    if not ruta_excel or not bandeja_seleccionada.get():
        messagebox.showwarning("Advertencia", "Primero selecciona un archivo y una hoja.")
        return
    
    try:
        df = pd.read_excel(ruta_excel, sheet_name=bandeja_seleccionada.get())
        ventana_ver = tk.Toplevel(ventana)
        ventana_ver.title("Inventario")
        
        tree = ttk.Treeview(ventana_ver, columns=list(df.columns), show="headings")
        for col in df.columns:
            tree.heading(col, text=col)
            tree.column(col, minwidth=0, width=100)
        
        for row in df.itertuples(index=False):
            tree.insert("", tk.END, values=row)
        
        tree.pack(expand=True, fill=tk.BOTH)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo mostrar el inventario: {e}")

# Crear etiquetas y campos de entrada en la ventana
for idx, (columna, variable) in enumerate(campos.items(), start=3):
    tk.Label(ventana, text=columna, bg="#F0E6F7", fg="#333", font=("Arial", 10, "bold")).grid(row=idx, column=0, padx=10, pady=5, sticky="w")
    tk.Entry(ventana, textvariable=variable, width=30).grid(row=idx, column=1, padx=10, pady=5)

# Botones para añadir, restar productos y ver el inventario
boton_añadir = tk.Button(ventana, text="Añadir Producto", command=lambda: actualizar_stock("añadir"), bg="#F9C8D1", fg="white", font=("Arial", 12, "bold"))
boton_añadir.grid(row=len(campos) + 4, column=0, columnspan=2, pady=10)

boton_restar = tk.Button(ventana, text="Restar Producto", command=lambda: actualizar_stock("restar"), bg="#D17A5B", fg="white", font=("Arial", 12, "bold"))
boton_restar.grid(row=len(campos) + 5, column=0, columnspan=2, pady=10)

boton_ver_articulos = tk.Button(ventana, text="Ver Artículos", command=ver_articulos, bg="#F3A5F3", fg="white", font=("Arial", 12, "bold"))
boton_ver_articulos.grid(row=len(campos) + 6, column=0, columnspan=2, pady=10)

# Ejecutar la ventana principal
ventana.mainloop()






